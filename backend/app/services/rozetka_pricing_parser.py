"""Rozetka pricing/commission Excel parser.

Parses the Rozetka commission XLSX file (sheet "Тариф") into structured
pricing rules.  The file contains Rozetka's own category IDs which match
the `channel_external_categories.external_id` column.

Commission is Rozetka's share — the selling price is:
    price = cost / (1 - commission_decimal)
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

logger = logging.getLogger("rozetka_pricing.parser")

# If the brand column contains "-", treat as "any brand"
BRAND_ANY = "-"

# If the price range column contains "-", treat as "any price"
RANGE_ANY = "-"

# Price range pattern: "min-max" or "min-999999999"
RANGE_RE = re.compile(r"^(\d+)-(\d+)$")

# Maximum valid commission percentage
MAX_COMMISSION = 99.0


@dataclass
class PricingRuleRow:
    """One parsed row from the Rozetka commission spreadsheet."""
    row_number: int = 0
    external_category_id: str = ""
    category_name: str = ""
    commission_percent: float = 0.0
    brand: Optional[str] = None
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    errors: list[str] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return len(self.errors) == 0


@dataclass
class ParseResult:
    """Result of parsing the pricing file."""
    rows: list[PricingRuleRow] = field(default_factory=list)
    invalid: list[PricingRuleRow] = field(default_factory=list)
    duplicates: int = 0
    total_rows: int = 0
    unique_categories: set[str] = field(default_factory=set)
    errors: list[str] = field(default_factory=list)


def parse_rozetka_pricing(filepath: str) -> ParseResult:
    """Parse the Rozetka commission XLSX file.

    Expected columns (Row 1 headers):
      1. ID категорії — Rozetka category ID (integer)
      2. Категорія — category name
      3. Бренд — brand (or "-" for any)
      4. Діапазон цін — price range "min-max" (or "-" for any)
      5. Відсоток комісії — commission percentage (float)

    Returns a ParseResult with valid rows, invalid rows, and stats.
    """
    result = ParseResult()

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        result.errors.append(f"Не вдалося відкрити файл: {e}")
        return result

    if "Тариф" not in wb.sheetnames:
        result.errors.append("Склад відсутній: 'Тариф'")
        return result

    ws = wb["Тариф"]
    if ws.max_row is None or ws.max_row < 2:
        result.errors.append("Файл не містить даних")
        return result

    seen_keys: set[tuple] = set()
    last_cat_id: Optional[str] = None
    last_cat_name: Optional[str] = None

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        result.total_rows += 1
        cat_id, cat_name, brand, price_range, commission = row

        rule = PricingRuleRow(row_number=i)

        # Category ID: inherit from previous row when merged cells
        if cat_id is not None:
            try:
                rule.external_category_id = str(int(cat_id))
                last_cat_id = rule.external_category_id
                last_cat_name = str(cat_name or "").strip()
                rule.category_name = last_cat_name
            except (ValueError, TypeError):
                rule.errors.append(f"Невірний ID категорії: {cat_id!r}")
                result.invalid.append(rule)
                continue
        else:
            if last_cat_id is None:
                continue
            rule.external_category_id = last_cat_id
            rule.category_name = last_cat_name or ""

        # Commission
        if commission is None:
            rule.errors.append("Відсутній відсоток комісії")
        else:
            try:
                rule.commission_percent = float(commission)
                if rule.commission_percent < 0 or rule.commission_percent > MAX_COMMISSION:
                    rule.errors.append(f"Комісія {rule.commission_percent}% поза допустимим діапазоном")
            except (ValueError, TypeError):
                rule.errors.append(f"Невірний відсоток комісії: {commission!r}")

        # Brand
        if brand and str(brand).strip() not in ("", BRAND_ANY):
            rule.brand = str(brand).strip()

        # Price range
        if price_range and str(price_range).strip() not in ("", RANGE_ANY):
            m = RANGE_RE.match(str(price_range).strip())
            if m:
                rule.price_min = int(m.group(1))
                rule.price_max = int(m.group(2))
            else:
                rule.errors.append(f"Невірний діапазон цін: {price_range!r}")

        if rule.valid:
            key = (rule.external_category_id, rule.brand, rule.price_min, rule.price_max)
            if key in seen_keys:
                result.duplicates += 1
                rule.errors.append("Дублікат правила")
                result.invalid.append(rule)
            else:
                seen_keys.add(key)
                result.rows.append(rule)
                result.unique_categories.add(rule.external_category_id)
        else:
            result.invalid.append(rule)

    return result